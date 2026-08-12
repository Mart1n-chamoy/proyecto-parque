"""
apps/dashboard/views.py

Views para el frontend web del sistema de cobranzas.
Agregar en settings.py INSTALLED_APPS: 'apps.dashboard'
Agregar en proyecto_cobranza/urls.py:
    path('', include('apps.dashboard.urls')),
"""

import io
import logging
import pandas as pd
from django.db import models
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views import View
from django.http import JsonResponse
from django.contrib.auth.mixins import LoginRequiredMixin

from apps.calls.models import CallBatch, Call
from apps.clients.models import Client
from apps.calls.tasks import process_call_batch, process_whatsapp_batch

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = {"phone_number", "name", "amount"}


class DashboardView(LoginRequiredMixin, View):
    """GET / — pantalla principal con stats y gráficos"""

    def get(self, request):
        from django.db.models import Avg, Count
        from django.utils import timezone
        from django.db.models.functions import TruncDate
        import datetime, json

        period = int(request.GET.get("period", 30))
        since  = timezone.now() - datetime.timedelta(days=period)

        batches         = CallBatch.objects.order_by("-created_at")[:10]
        total_calls     = Call.objects.count()
        completed_calls = Call.objects.filter(status="completed").count()
        failed_calls    = Call.objects.filter(status="failed").count()
        voicemail_calls = Call.objects.filter(outcome="voicemail").count()
        failed_only     = max(failed_calls - voicemail_calls, 0)
        success_rate    = round(completed_calls / total_calls * 100) if total_calls else 0

        dur_qs = (
            CallBatch.objects
            .annotate(avg_dur=Avg("calls__duration"))
            .filter(avg_dur__isnull=False)
            .order_by("-created_at")[:8]
        )
        chart_duracion = {
            "labels": json.dumps([b.name for b in dur_qs][::-1]),
            "data":   json.dumps([round(b.avg_dur or 0, 1) for b in dur_qs][::-1]),
        }

        dias_qs = (
            Call.objects
            .filter(created_at__gte=since)
            .annotate(day=TruncDate("created_at"))
            .values("day")
            .annotate(total=Count("id"))
            .order_by("day")
        )
        chart_dias = {
            "labels": json.dumps([str(d["day"]) for d in dias_qs]),
            "data":   json.dumps([d["total"] for d in dias_qs]),
        }

        return render(request, "dashboard/index.html", {
            "batches":        batches,
            "period":         period,
            "chart_duracion": chart_duracion,
            "chart_dias":     chart_dias,
            "stats": {
                "total_batches":     CallBatch.objects.count(),
                "total_calls":       total_calls,
                "completed_calls":   completed_calls,
                "failed_calls":      failed_calls,
                "voicemail_calls":   voicemail_calls,
                "failed_only_calls": failed_only,
                "success_rate":      success_rate,
            },
        })


class CampaignListView(LoginRequiredMixin, View):
    """GET /campaigns/ — lista completa de campañas (paginada, con filtros)"""

    def get(self, request):
        from django.core.paginator import Paginator

        qs = CallBatch.objects.order_by("-created_at")

        status_filter = request.GET.get("status", "")
        if status_filter:
            qs = qs.filter(status=status_filter)

        channel_filter = request.GET.get("channel", "")
        if channel_filter:
            qs = qs.filter(channel=channel_filter)

        search = request.GET.get("q", "").strip()
        if search:
            qs = qs.filter(name__icontains=search)

        paginator = Paginator(qs, 20)
        page_obj = paginator.get_page(request.GET.get("page"))

        return render(request, "dashboard/campaign_list.html", {
            "page_obj": page_obj,
            "status_filter": status_filter,
            "channel_filter": channel_filter,
            "search": search,
            "status_choices": CallBatch.STATUS_CHOICES,
            "channel_choices": CallBatch.CHANNEL_CHOICES,
        })


class CallListView(LoginRequiredMixin, View):
    """GET /calls/ — lista completa de llamadas/mensajes individuales (paginada, con filtros)"""

    def get(self, request):
        from django.core.paginator import Paginator

        qs = Call.objects.select_related("client", "batch").order_by("-created_at")

        status_filter = request.GET.get("status", "")
        if status_filter:
            qs = qs.filter(status=status_filter)

        channel_filter = request.GET.get("channel", "")
        if channel_filter:
            qs = qs.filter(channel=channel_filter)

        search = request.GET.get("q", "").strip()
        if search:
            qs = qs.filter(
                models.Q(client__first_name__icontains=search) |
                models.Q(client__last_name__icontains=search) |
                models.Q(client__phone__icontains=search)
            )

        paginator = Paginator(qs, 30)
        page_obj = paginator.get_page(request.GET.get("page"))

        return render(request, "dashboard/call_list.html", {
            "page_obj": page_obj,
            "status_filter": status_filter,
            "channel_filter": channel_filter,
            "search": search,
            "status_choices": Call.STATUS_CHOICES,
            "channel_choices": CallBatch.CHANNEL_CHOICES,
        })


class CampaignNewView(LoginRequiredMixin,View):
    """GET/POST /campaigns/new/ — crear campaña y subir CSV"""

    def get(self, request):
        return render(request, "dashboard/campaign_new.html")

    def post(self, request):
        name    = request.POST.get("name", "").strip()
        action  = request.POST.get("action", "save")
        file    = request.FILES.get("file")
        channel = request.POST.get("channel", "call")
        whatsapp_template_name     = request.POST.get("whatsapp_template_name", "").strip()
        whatsapp_template_language = request.POST.get("whatsapp_template_language", "es").strip() or "es"

        if not name:
            messages.error(request, "El nombre de la campaña es obligatorio.")
            return render(request, "dashboard/campaign_new.html")

        if not file:
            messages.error(request, "Debés subir un archivo CSV o Excel.")
            return render(request, "dashboard/campaign_new.html")

        if channel == "whatsapp" and not whatsapp_template_name:
            messages.error(request, "Elegí un template de WhatsApp aprobado en Meta.")
            return render(request, "dashboard/campaign_new.html")

        # Leer archivo
        try:
            if file.name.endswith(".csv"):
                df = pd.read_csv(io.BytesIO(file.read()))
            else:
                df = pd.read_excel(io.BytesIO(file.read()))
        except Exception as e:
            messages.error(request, f"No se pudo leer el archivo: {e}")
            return render(request, "dashboard/campaign_new.html")

        # Validar columnas
        missing = REQUIRED_COLUMNS - set(df.columns.str.lower())
        if missing:
            messages.error(request, f"Faltan columnas en el archivo: {', '.join(missing)}")
            return render(request, "dashboard/campaign_new.html")

        df.columns = df.columns.str.lower()

        # Crear el lote
        batch = CallBatch.objects.create(
            name=name,
            status="pending",
            total_clients=len(df),
            channel=channel,
            whatsapp_template_name=whatsapp_template_name if channel == "whatsapp" else None,
            whatsapp_template_language=whatsapp_template_language if channel == "whatsapp" else None,
        )

        # Crear clientes y llamadas
        calls_created = 0
        for _, row in df.iterrows():
            phone = str(row["phone_number"]).strip()
            if not phone:
                continue

            # Obtener o crear cliente
            name_parts = str(row.get("name", "")).strip().split(" ", 1)
            first_name = name_parts[0] if name_parts else ""
            last_name  = name_parts[1] if len(name_parts) > 1 else ""
            
            client, _ = Client.objects.get_or_create(
                phone=phone,
                defaults={
                            "first_name": first_name,
                            "last_name":  last_name,
                            "debt_amount": row.get("amount", 0),
                        }
            )
            # Actualizar deuda si el cliente ya existía
            client.debt_amount = row.get("amount", client.debt_amount)
            client.save(update_fields=["debt_amount"])

            Call.objects.create(
                batch=batch,
                client=client,
                status="pending",
                channel=channel,
            )
            calls_created += 1

        batch.total_clients = calls_created
        batch.save(update_fields=["total_clients"])

        canal_label = "WhatsApp" if channel == "whatsapp" else "llamadas"
        messages.success(request, f"Campaña creada con {calls_created} clientes ({canal_label}).")

        # Si eligieron lanzar ahora, encolar la tarea Celery según el canal
        if action == "launch":
            if channel == "whatsapp":
                process_whatsapp_batch.delay(batch.id)
                messages.success(request, "Los mensajes de WhatsApp se están enviando en background.")
            else:
                process_call_batch.delay(batch.id)
                messages.success(request, "Las llamadas se están iniciando en background.")

        return redirect(f"/campaigns/{batch.id}/")


class CampaignDetailView(LoginRequiredMixin, View):
    """GET /campaigns/{id}/ — detalle de campaña con llamadas"""

    def get(self, request, pk):
        batch = get_object_or_404(CallBatch, pk=pk)
        calls = Call.objects.filter(batch=batch).select_related("client").order_by("id")

        completed = calls.filter(status="completed").count()
        failed    = calls.filter(status="failed").count()
        pending   = calls.filter(status__in=["pending", "in_progress"]).count()

        return render(request, "dashboard/campaign_detail.html", {
            "batch": batch,
            "calls": calls,
            "stats": {
                "completed": completed,
                "failed":    failed,
                "pending":   pending,
            },
        })


class CampaignLaunchView(LoginRequiredMixin, View):
    """POST /campaigns/{id}/launch/ — lanzar lote a ElevenLabs"""

    def post(self, request, pk):
        batch = get_object_or_404(CallBatch, pk=pk)

        if batch.status not in ("pending", "failed"):
            messages.error(request, f"El lote no se puede lanzar en estado '{batch.status}'.")
            return redirect(f"/campaigns/{batch.id}/")

        if batch.channel == "whatsapp":
            process_whatsapp_batch.delay(batch.id)
            messages.success(request, "Mensajes de WhatsApp iniciados. Los resultados se actualizarán automáticamente.")
        else:
            process_call_batch.delay(batch.id)
            messages.success(request, "Llamadas iniciadas. Los resultados se actualizarán automáticamente.")
        return redirect(f"/campaigns/{batch.id}/")


class CampaignStatusView(LoginRequiredMixin, View):
    """GET /campaigns/{id}/status/ — fragmento HTMX con tabla actualizada"""

    def get(self, request, pk):
        batch = get_object_or_404(CallBatch, pk=pk)
        calls = Call.objects.filter(batch=batch).select_related("client").order_by("id")

        completed = calls.filter(status="completed").count()
        failed    = calls.filter(status="failed").count()
        pending   = calls.filter(status__in=["pending", "in_progress"]).count()

        return render(request, "dashboard/campaign_detail.html", {
            "batch": batch,
            "calls": calls,
            "stats": {
                "completed": completed,
                "failed":    failed,
                "pending":   pending,
            },
        })

"""
Agregar en apps/dashboard/views.py — views de autenticación para el panel web.

Usa sesiones Django (no JWT) para el frontend de templates.
JWT sigue funcionando para la API REST como estaba.
"""

from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect
from django.views import View


class DashboardLoginView(View):
    """GET/POST /login/"""

    def get(self, request):
        if request.user.is_authenticated:
            return redirect("/")
        return render(request, "dashboard/login.html")

    def post(self, request):
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            next_url = request.GET.get("next", "/")
            return redirect(next_url)

        return render(request, "dashboard/login.html", {
            "error": "Usuario o contraseña incorrectos."
        })


class DashboardLogoutView(View):
    """POST /logout/"""

    def post(self, request):
        logout(request)
        return redirect("/login/")

import csv
import io


class CampaignExportView(LoginRequiredMixin, View):
    """
    GET /campaigns/{id}/export/?format=csv
    GET /campaigns/{id}/export/?format=xlsx
    """

    def get(self, request, pk):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from apps.calls.models import CallBatch, Call

        batch = get_object_or_404(CallBatch, pk=pk)
        calls = Call.objects.filter(batch=batch).select_related("client").order_by("id")

        fmt = request.GET.get("format", "csv")

        headers = ["Cliente", "Teléfono", "Deuda", "Moneda", "Estado", "Resultado", "Duración (seg)", "Transcripción"]
        rows = []
        for call in calls:
            rows.append([
                (call.client.first_name + " " + (call.client.last_name or "")).strip(),
                call.client.phone,
                str(call.client.debt_amount or ""),
                getattr(call.client, "currency", "ARS"),
                call.get_status_display(),
                call.outcome or "",
                str(call.duration or ""),
                call.transcript or "",
            ])

        if fmt == "xlsx":
            return self._export_xlsx(batch, headers, rows)
        return self._export_csv(batch, headers, rows)

    def _export_csv(self, batch, headers, rows):
        from django.http import HttpResponse
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="campana_{batch.id}.csv"'
        response.write("\ufeff")
        writer = csv.writer(response)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        return response

    def _export_xlsx(self, batch, headers, rows):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Campaña {batch.id}"

        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        col_widths = [25, 18, 12, 10, 14, 14, 16, 60]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="campana_{batch.id}.xlsx"'
        return response
